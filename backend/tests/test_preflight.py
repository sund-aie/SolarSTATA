"""Pre-flight inspection of a staged xlsx upload.

Canonical case: the workbook with the header on row 5 — that is the
exact scenario the old hardcoded row-1 default in the HeaderRowPicker
silently got wrong. The detected_header_row + notes_rows pair must
land at 5 and [1, 2, 3, 4] respectively.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
from fastapi.testclient import TestClient

from solarstata.io import preflight_xlsx


def _tidy_long_workbook(path: Path) -> None:
    """Title rows + blanks + header on row 5 — the real-world case."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tidy"
    ws.append(["TIDY LONG FORMAT"])           # row 1 — title
    ws.append(["Anonymized clinical extract"])  # row 2 — subtitle
    ws.append([])                              # row 3 — blank
    ws.append([])                              # row 4 — blank
    ws.append(["patient_id", "age", "sex"])    # row 5 — header
    ws.append([1001, 23, "F"])
    ws.append([1002, 45, "M"])
    ws.append([1003, 31, "F"])
    wb.save(str(path))


def _happy_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["age", "sex", "weight"])
    ws.append([23, "F", 70.5])
    ws.append([45, "M", 82.0])
    ws.append([31, "F", 65.0])
    wb.save(str(path))


def _merged_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["title spans cells"])
    ws.merge_cells("A1:C1")
    ws.append([])
    ws.append(["id", "name", "score"])
    ws.append([1, "Alice", 90])
    ws.append([2, "Bob", 85])
    wb.save(str(path))


def _hidden_rows_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["id", "value"])
    ws.append([1, 10])
    ws.append([2, 20])
    ws.append([3, 30])
    ws.row_dimensions[3].hidden = True
    wb.save(str(path))


# ---------------------------------------------------------------------------
# Direct preflight_xlsx tests
# ---------------------------------------------------------------------------

def test_preflight_detects_tidy_long_header_on_row_5(tmp_path: Path) -> None:
    """Canonical regression: header on row 5, notes rows are 1 through 4.

    Both assertions matter — the row count is what the auto-default
    keys off, and notes_rows is what the UI strip uses to tell the
    user "rows 1 to 4 will be skipped" regardless of whether some of
    those rows were blank (rows 3 and 4 here) or had content.
    """
    p = tmp_path / "tidy.xlsx"
    _tidy_long_workbook(p)

    result = preflight_xlsx(p)
    assert result["detected_header_row"] == 5
    assert result["notes_rows"] == [1, 2, 3, 4]
    assert result["header_cells"] == ["patient_id", "age", "sex"]


def test_preflight_happy_path_header_on_row_1(tmp_path: Path) -> None:
    p = tmp_path / "happy.xlsx"
    _happy_workbook(p)

    result = preflight_xlsx(p)
    assert result["detected_header_row"] == 1
    assert result["notes_rows"] == []
    assert result["header_cells"] == ["age", "sex", "weight"]


def test_preflight_column_kinds_split(tmp_path: Path) -> None:
    """patient_id → identifier, age → numeric, sex → categorical."""
    p = tmp_path / "tidy.xlsx"
    _tidy_long_workbook(p)
    kinds = preflight_xlsx(p)["column_kinds"]
    assert kinds["identifier"] == 1
    assert kinds["numeric"] == 1
    # sex has 2 unique values across 3 rows → categorical bucket
    assert kinds["categorical"] == 1
    assert kinds["string"] == 0


def test_preflight_flags_merged_cells(tmp_path: Path) -> None:
    p = tmp_path / "merged.xlsx"
    _merged_workbook(p)
    issues = preflight_xlsx(p)["cell_issues"]
    assert issues["merged_cells"] == 1


def test_preflight_flags_hidden_rows(tmp_path: Path) -> None:
    p = tmp_path / "hidden.xlsx"
    _hidden_rows_workbook(p)
    issues = preflight_xlsx(p)["cell_issues"]
    assert issues["hidden_rows"] == 1


def test_preflight_unknown_sheet_raises(tmp_path: Path) -> None:
    p = tmp_path / "happy.xlsx"
    _happy_workbook(p)
    try:
        preflight_xlsx(p, sheet="DoesNotExist")
    except KeyError as e:
        assert "DoesNotExist" in str(e)
    else:
        raise AssertionError("expected KeyError for unknown sheet")


# ---------------------------------------------------------------------------
# Route-level tests (staging round-trip + non-xlsx rejection)
# ---------------------------------------------------------------------------

def test_preflight_route_against_staged_upload(client: TestClient, tmp_path: Path) -> None:
    p = tmp_path / "tidy.xlsx"
    _tidy_long_workbook(p)

    with p.open("rb") as fh:
        stage = client.post(
            "/api/data/upload",
            files={"file": ("tidy.xlsx", fh, "application/octet-stream")},
        ).json()
    assert stage["requires_choice"] is True

    resp = client.post(
        "/api/data/preflight",
        json={"file_id": stage["file_id"], "sheet": "Tidy"},
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["detected_header_row"] == 5
    assert body["notes_rows"] == [1, 2, 3, 4]


def test_preflight_unknown_file_id_returns_404(client: TestClient) -> None:
    resp = client.post(
        "/api/data/preflight",
        json={"file_id": "nonexistent", "sheet": "x"},
    )
    assert resp.status_code == 404


def test_preflight_default_sheet_is_first(client: TestClient, tmp_path: Path) -> None:
    """Omitting `sheet` should use the first sheet of the workbook."""
    p = tmp_path / "tidy.xlsx"
    _tidy_long_workbook(p)
    with p.open("rb") as fh:
        stage = client.post(
            "/api/data/upload",
            files={"file": ("tidy.xlsx", fh, "application/octet-stream")},
        ).json()

    resp = client.post("/api/data/preflight", json={"file_id": stage["file_id"]})
    assert resp.status_code == 200, resp.text
    assert resp.json()["sheet"] == "Tidy"
