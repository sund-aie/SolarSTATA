"""Pre-flight inspection of a staged xlsx upload.

Reads back the workbook with openpyxl to surface:
- the most-likely header row (better than the old always-row-1 default
  in the HeaderRowPicker — the canonical break was a file with the
  header on row 5 under four note/blank rows)
- every row above the detected header — both content rows and blanks
  — so the UI can say plainly "rows 1 to N-1 will be skipped"
- a column-kind summary derived from the data rows directly below the
  header (numeric / categorical / identifier / string)
- structural problems the user should know about before committing
  (merged cells, hidden rows, hidden columns)

Pure function. Does not touch the staging store or the session — the
caller resolves file_id → path before calling in.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def preflight_xlsx(
    path: str | Path,
    *,
    sheet: str | None = None,
    scan_rows: int = 30,
) -> dict[str, Any]:
    """Inspect the given workbook + sheet without committing the dataset.

    `scan_rows` caps how far we look from the top of the sheet — the
    header heuristic only needs a finite head, and a small cap keeps
    the worst case bounded on giant workbooks.
    """
    wb = load_workbook(filename=str(path), data_only=True)
    try:
        chosen = sheet if sheet is not None else wb.sheetnames[0]
        if chosen not in wb.sheetnames:
            raise KeyError(f"sheet {chosen!r} not in workbook")
        ws = wb[chosen]

        rows: list[list[Any]] = []
        for i, row in enumerate(ws.iter_rows(max_row=scan_rows, values_only=True)):
            rows.append(list(row))

        detected = _detect_header_row(rows)
        # Every row above the header — including blanks. The UI uses
        # this to say "rows 1 to N-1 will be skipped", and the user
        # cares about the *position* of those rows, not their content.
        notes_rows = list(range(1, detected))

        header_cells = (
            [_stringify(c) for c in rows[detected - 1]]
            if detected - 1 < len(rows)
            else []
        )
        # Trim trailing empty header cells so the column-kind sniff
        # doesn't keep counting phantom columns past the real width.
        while header_cells and header_cells[-1] == "":
            header_cells.pop()

        data_rows = rows[detected:]
        column_kinds = _sniff_column_kinds(header_cells, data_rows)
        cell_issues = _cell_issues(ws)

        n_rows_total = int(ws.max_row or 0)
        n_rows_after_header = max(0, n_rows_total - detected)

        return {
            "sheet": chosen,
            "detected_header_row": detected,
            "notes_rows": notes_rows,
            "header_cells": header_cells,
            "column_kinds": column_kinds,
            "cell_issues": cell_issues,
            "n_rows_after_header": n_rows_after_header,
        }
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Header-row heuristic
# ---------------------------------------------------------------------------

def _detect_header_row(rows: list[list[Any]]) -> int:
    """Return the 1-based row index that looks most like a header.

    Walk top-down. A row is a header candidate when:
      - it has at least 2 populated cells (single-cell rows are notes)
      - at least 50% of its populated cells are text (numbers in row
        positions are data, not labels)
      - the row directly below it has either a numeric cell or at
        least as many populated cells as the candidate (i.e. it
        looks like a data row, not another note)

    No candidate found → fall back to row 1 so the picker behaves
    exactly as before.
    """
    for r in range(len(rows) - 1):
        populated = _populated(rows[r])
        if len(populated) < 2:
            continue
        text_share = sum(1 for c in populated if _is_text(c)) / len(populated)
        if text_share < 0.5:
            continue
        below = _populated(rows[r + 1])
        if not below:
            continue
        has_numeric = any(_is_numeric(c) for c in below)
        if has_numeric or len(below) >= len(populated):
            return r + 1
    return 1


def _populated(row: list[Any]) -> list[Any]:
    return [c for c in row if c not in (None, "")]


def _is_text(v: Any) -> bool:
    return isinstance(v, str) and v.strip() != ""


def _is_numeric(v: Any) -> bool:
    return isinstance(v, (int, float)) and not isinstance(v, bool)


# ---------------------------------------------------------------------------
# Column-kind sniff
# ---------------------------------------------------------------------------

def _sniff_column_kinds(
    header_cells: list[str],
    data_rows: list[list[Any]],
) -> dict[str, int]:
    """Quick classification of each column based on the preview data.

    This is *informational* — the real Frame-level classification
    runs after the dataset commits. The goal here is to give the
    user a sense of shape ("8 numeric, 3 categorical, 1 id") in the
    pre-flight strip, not to replace the downstream classifier.
    """
    kinds = {"numeric": 0, "categorical": 0, "identifier": 0, "string": 0}
    for col_idx, name in enumerate(header_cells):
        vals = [r[col_idx] if col_idx < len(r) else None for r in data_rows]
        vals = [v for v in vals if v not in (None, "")]
        if not vals:
            kinds["string"] += 1
            continue
        lname = name.lower()
        all_numeric = all(_is_numeric(v) for v in vals)
        all_int = all(isinstance(v, int) and not isinstance(v, bool) for v in vals)
        if all_numeric and all_int and ("id" in lname):
            kinds["identifier"] += 1
        elif all_numeric:
            kinds["numeric"] += 1
        elif _looks_categorical(vals):
            kinds["categorical"] += 1
        else:
            kinds["string"] += 1
    return kinds


def _looks_categorical(vals: list[Any]) -> bool:
    """Categorical when unique count is small both absolutely and
    relative to the sample. ≤10 unique and at most half the sample."""
    n_unique = len(set(vals))
    return n_unique <= max(2, len(vals) // 2) and n_unique <= 10


# ---------------------------------------------------------------------------
# Cell-structure issues
# ---------------------------------------------------------------------------

def _cell_issues(ws) -> dict[str, int]:
    """Count merged cells and hidden rows/columns from openpyxl metadata."""
    merged = len(ws.merged_cells.ranges) if hasattr(ws, "merged_cells") else 0
    hidden_rows = sum(1 for rd in ws.row_dimensions.values() if rd.hidden)
    hidden_cols = sum(1 for cd in ws.column_dimensions.values() if cd.hidden)
    return {
        "merged_cells": int(merged),
        "hidden_rows": int(hidden_rows),
        "hidden_cols": int(hidden_cols),
    }


def _stringify(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)
