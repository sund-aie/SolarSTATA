"""Dataset readers.

Returns a Frame with as much metadata preserved as the source format
allows. .dta retains variable labels and value labels via pyreadstat;
csv/xlsx/parquet have no native equivalent so those fields stay empty.

`read_dataset` accepts optional `sheet` and `header_row` for xlsx files
so the caller can disambiguate multi-sheet workbooks and skip
title/subtitle rows above the actual headers (the "TIDY LONG FORMAT"
case). Both default to the first sheet and row 1.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
import pyreadstat

from openpyxl import load_workbook

from ..session.models import Frame

SUPPORTED_FORMATS = ("csv", "tsv", "xlsx", "xls", "dta", "parquet")


def sniff_format(filename: str) -> str:
    suffix = Path(filename).suffix.lower().lstrip(".")
    if suffix in ("csv",):
        return "csv"
    if suffix in ("tsv", "txt"):
        return "tsv"
    if suffix in ("xlsx", "xls"):
        return "xlsx"
    if suffix in ("dta",):
        return "dta"
    if suffix in ("parquet", "pq"):
        return "parquet"
    raise ValueError(
        f"Unsupported file format: .{suffix}. Supported: {', '.join(SUPPORTED_FORMATS)}"
    )


def read_dataset(
    path: str | Path,
    *,
    name: str = "default",
    sheet: str | None = None,
    header_row: int = 1,
) -> Frame:
    """Read a file off disk into a Frame, preserving labels where possible.

    `header_row` is 1-based to match Stata's usual convention; pandas wants
    0-based, so we subtract 1 internally. Rows above the header row are
    discarded; fully-blank rows that bleed in between header and data are
    dropped after the read.
    """
    path = Path(path)
    fmt = sniff_format(path.name)

    column_labels: dict[str, str] = {}
    value_labels: dict[str, dict] = {}

    if fmt == "csv":
        df = pd.read_csv(path, header=header_row - 1)
    elif fmt == "tsv":
        df = pd.read_csv(path, sep="\t", header=header_row - 1)
    elif fmt == "xlsx":
        df = pd.read_excel(path, sheet_name=sheet if sheet is not None else 0, header=header_row - 1)
        # When a header row sits below blank/title rows, pandas can carry
        # those blank rows in as all-NaN. Drop them.
        df = df.dropna(how="all").reset_index(drop=True)
    elif fmt == "parquet":
        df = pd.read_parquet(path)
    elif fmt == "dta":
        df, meta = pyreadstat.read_dta(path)
        column_labels = dict(zip(meta.column_names, meta.column_labels)) if meta.column_labels else {}
        column_labels = {k: v for k, v in column_labels.items() if v}
        value_labels = dict(meta.variable_value_labels or {})
    else:  # pragma: no cover — sniff_format already rejected unknowns
        raise ValueError(f"Unhandled format: {fmt}")

    df.columns = [str(c) for c in df.columns]
    df = _dedupe_columns(df)

    storage_types = {col: _stata_storage_type(df[col]) for col in df.columns}

    return Frame(
        name=name,
        df=df,
        column_labels=column_labels,
        value_labels=value_labels,
        storage_types=storage_types,
        source_filename=path.name,
    )


def list_xlsx_sheets(path: str | Path, *, n_preview_rows: int = 10) -> list[dict[str, Any]]:
    """Inspect a workbook without committing the dataset to memory.

    Returns one dict per sheet with name, dimensions, and the raw first N
    rows so the frontend can offer both a sheet picker and a header-row
    picker before the user commits to the upload.
    """
    wb = load_workbook(filename=str(path), read_only=True, data_only=True)
    try:
        out: list[dict[str, Any]] = []
        for name in wb.sheetnames:
            ws = wb[name]
            preview: list[list[str]] = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i >= n_preview_rows:
                    break
                preview.append([_stringify_cell(v) for v in row])
            out.append({
                "name": name,
                "n_rows": int(ws.max_row or 0),
                "n_cols": int(ws.max_column or 0),
                "preview_rows": preview,
            })
        return out
    finally:
        wb.close()


def _stringify_cell(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def _dedupe_columns(df: pd.DataFrame) -> pd.DataFrame:
    """If a CSV/Excel has repeated headers, suffix later ones with _1, _2, ..."""
    seen: dict[str, int] = {}
    new_cols = []
    for col in df.columns:
        if col not in seen:
            seen[col] = 0
            new_cols.append(col)
        else:
            seen[col] += 1
            new_cols.append(f"{col}_{seen[col]}")
    df.columns = new_cols
    return df


def _stata_storage_type(series: pd.Series) -> str:
    """Map a pandas dtype to its Stata storage-type name.

    This is an approximation — Stata's compress would pick the smallest
    integer type that fits, but for Phase 1 we just classify int vs
    float vs string. Phase 1.1 will add a real `compress` routine that
    inspects observed value ranges.
    """
    dtype = series.dtype
    if pd.api.types.is_integer_dtype(dtype):
        return "long"
    if pd.api.types.is_float_dtype(dtype):
        return "double"
    if pd.api.types.is_bool_dtype(dtype):
        return "byte"
    return "str"
